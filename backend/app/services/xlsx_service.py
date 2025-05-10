"""
XLSX export service for generating Excel files from extraction results.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any, Union
from uuid import UUID
import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from ..models import ExtractionJob, ExtractionResult, ProcessingStatus, XLSXExport

# Configure logging
logger = logging.getLogger(__name__)

# Constants
EXCEL_DIR = Path("excel_exports")
EXCEL_DIR.mkdir(exist_ok=True)


class XLSXExportService:
    """Service for generating Excel files from extraction results."""
    
    @staticmethod
    def sanitize_data(content: Dict[str, Any]) -> Dict[str, Any]:
        """
        Sanitize and clean data to make it Excel-compatible.
        
        Args:
            content: The content dictionary to sanitize
            
        Returns:
            Dict[str, Any]: Sanitized content dictionary
        """
        if not content:
            return {}
        
        # Make a copy to avoid modifying the original
        clean_content = {}
        
        # Process each key in the content dictionary
        for key, value in content.items():
            # Handle nested dictionaries
            if isinstance(value, dict):
                clean_content[key] = XLSXExportService.sanitize_data(value)
            
            # Handle lists - this is where we often have issues
            elif isinstance(value, list):
                # Check if it's a list of dictionaries
                if all(isinstance(item, dict) for item in value):
                    # Convert list of dicts to a simpler format
                    clean_list = []
                    for item in value:
                        # Recursively sanitize each dict in the list
                        clean_item = XLSXExportService.sanitize_data(item)
                        clean_list.append(clean_item)
                    clean_content[key] = clean_list
                else:
                    # For simple lists, just convert to string
                    try:
                        clean_content[key] = json.dumps(value)
                    except:
                        clean_content[key] = str(value)
            
            # Handle other types
            else:
                # Convert to string if it's not a basic type
                if not isinstance(value, (str, int, float, bool, type(None))):
                    try:
                        clean_content[key] = str(value)
                    except:
                        clean_content[key] = "Unconvertible data"
                else:
                    clean_content[key] = value
        
        return clean_content
    
    @staticmethod
    async def generate_xlsx(job_id: UUID, session: Session) -> XLSXExport:
        """
        Generate an XLSX file from extraction results.
        
        Args:
            job_id: The ID of the extraction job
            session: Database session
            
        Returns:
            XLSXExport: The XLSX export record
        """
        # Find the extraction job
        job = session.exec(
            select(ExtractionJob).where(ExtractionJob.id == job_id)
        ).one_or_none()
        
        if not job:
            raise ValueError(f"Extraction job with ID {job_id} not found")
        
        if job.status != ProcessingStatus.COMPLETED:
            raise ValueError(f"Extraction job is not completed (status: {job.status})")
        
        # Get all results for this job
        results = session.exec(
            select(ExtractionResult)
            .where(ExtractionResult.job_id == job_id)
            .order_by(ExtractionResult.page_number)
        ).all()
        
        if not results:
            raise ValueError(f"No extraction results found for job {job_id}")
        
        # Create workbook
        wb = Workbook()
        
        # Format main worksheet with extraction data
        ws = wb.active
        ws.title = "Extracted Data"
        
        # Add header
        ws.append(["Field Name", "Value", "Page", "Confidence"])
        
        # Style header
        for col_num in range(1, 5):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        
        # Add data
        row_num = 2
        for result in results:
            page_num = result.page_number
            
            # Process the content
            content = result.content
            if isinstance(content, dict):
                # Sanitize the content to ensure it's Excel-compatible
                try:
                    sanitized_content = XLSXExportService.sanitize_data(content)
                    
                    # Special handling for questions list
                    if "questions" in sanitized_content and isinstance(sanitized_content["questions"], list):
                        logger.info(f"Processing questions list with {len(sanitized_content['questions'])} items")
                        for i, question_data in enumerate(sanitized_content["questions"]):
                            if isinstance(question_data, dict):
                                question = question_data.get("question", f"Question {i+1}")
                                answer = question_data.get("answer", "")
                                confidence = question_data.get("confidence", 0)
                                
                                ws.append([question, answer, page_num, confidence])
                                
                                # Style the row
                                for col_num in range(1, 5):
                                    cell = ws.cell(row=row_num, column=col_num)
                                    cell.border = Border(
                                        left=Side(style="thin"),
                                        right=Side(style="thin"),
                                        top=Side(style="thin"),
                                        bottom=Side(style="thin")
                                    )
                                    
                                    # Highlight illegible values
                                    if col_num == 2 and answer == "[ILLEGIBLE]":
                                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                                
                                row_num += 1
                    
                    # Process other fields as before
                    for field_name, value in sanitized_content.items():
                        # Skip the questions list as it's already processed
                        if field_name == "questions":
                            continue
                        
                        # Convert complex values to strings
                        if isinstance(value, (dict, list)):
                            try:
                                value = json.dumps(value)
                            except:
                                value = str(value)
                        
                        ws.append([field_name, value, page_num, result.confidence_score or "N/A"])
                        
                        # Style the row
                        for col_num in range(1, 5):
                            cell = ws.cell(row=row_num, column=col_num)
                            cell.border = Border(
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                top=Side(style="thin"),
                                bottom=Side(style="thin")
                            )
                            
                            # Highlight illegible values
                            if col_num == 2 and value == "[ILLEGIBLE]":
                                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        
                        row_num += 1
                except Exception as e:
                    # If we encounter an error, log it and add an error row
                    logger.error(f"Error processing content: {str(e)}")
                    ws.append(["Error", f"Could not process content: {str(e)}", page_num, "N/A"])
                    row_num += 1
        
        # Auto-adjust column widths
        for col_num in range(1, 5):
            max_length = 0
            column = get_column_letter(col_num)
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
            ws.column_dimensions[column].width = adjusted_width
        
        # Add summary worksheet
        summary_ws = wb.create_sheet(title="Summary")
        
        # Add document info
        document = job.document
        summary_ws.append(["Document Information"])
        summary_ws.append(["Filename", document.filename])
        summary_ws.append(["Total Pages", document.total_pages])
        summary_ws.append(["Upload Date", document.uploaded_at.strftime("%Y-%m-%d %H:%M:%S")])
        summary_ws.append([""])
        
        # Add extraction info
        summary_ws.append(["Extraction Information"])
        summary_ws.append(["Model Used", job.model_name])
        summary_ws.append(["Pages Processed", job.pages_processed])
        summary_ws.append(["Started", job.started_at.strftime("%Y-%m-%d %H:%M:%S") if job.started_at else "N/A"])
        summary_ws.append(["Completed", job.completed_at.strftime("%Y-%m-%d %H:%M:%S") if job.completed_at else "N/A"])
        
        # Format the summary sheet
        for row_num in range(1, summary_ws.max_row + 1, 6):
            cell = summary_ws.cell(row=row_num, column=1)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Auto-adjust summary column widths
        for col_num in range(1, 3):
            max_length = 0
            column = get_column_letter(col_num)
            
            for row_num in range(1, summary_ws.max_row + 1):
                cell = summary_ws.cell(row=row_num, column=col_num)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
            summary_ws.column_dimensions[column].width = adjusted_width
        
        # Save to file
        filename = f"{document.filename.replace('.pdf', '')}_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = EXCEL_DIR / filename
        wb.save(file_path)
        
        # Get file size
        file_size = file_path.stat().st_size
        
        # Create XLSX export record
        xlsx_export = XLSXExport(
            job_id=job_id,
            filename=filename,
            file_size=file_size,
            local_path=str(file_path),
            generated_at=datetime.utcnow()
        )
        session.add(xlsx_export)
        session.commit()
        session.refresh(xlsx_export)
        
        return xlsx_export
    
    @staticmethod
    async def get_xlsx_file(export_id: UUID, session: Session) -> Optional[bytes]:
        """
        Get an XLSX file by export ID.
        
        Args:
            export_id: The ID of the XLSX export
            session: Database session
            
        Returns:
            Optional[bytes]: The XLSX file contents
        """
        # Find the XLSX export
        xlsx_export = session.exec(
            select(XLSXExport).where(XLSXExport.id == export_id)
        ).one_or_none()
        
        if not xlsx_export:
            return None
        
        # Check if file exists
        file_path = Path(xlsx_export.local_path)
        if not file_path.exists():
            return None
        
        # Read file contents
        with open(file_path, "rb") as f:
            content = f.read()
        
        return content
    
    @staticmethod
    def flatten_json_results(results: List[ExtractionResult]) -> pd.DataFrame:
        """
        Convert extraction results into a flat DataFrame.
        
        Args:
            results: List of extraction results
            
        Returns:
            pd.DataFrame: Flattened data
        """
        # Collect all data
        all_data = []
        
        for result in results:
            page_num = result.page_number
            
            # Process the content
            content = result.content
            if isinstance(content, dict):
                for field_name, value in content.items():
                    all_data.append({
                        "Field Name": field_name,
                        "Value": value,
                        "Page": page_num,
                        "Confidence": result.confidence_score or "N/A"
                    })
        
        # Convert to DataFrame
        return pd.DataFrame(all_data) 