from io import BytesIO
import base64
from typing import Dict, Any

import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from ..schemas import ExtractionResult


def to_xlsx_bytes(result: ExtractionResult) -> bytes:
    """Convert ExtractionResult to XLSX bytes (original implementation)"""
    data = [row.model_dump() for row in result.rows]
    df = pd.DataFrame(data)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    return buffer.getvalue()


class XLSXService:
    """Service for generating XLSX files from structured data"""
    
    def generate_from_data(self, data: Dict[str, Any]) -> bytes:
        """
        Generate an XLSX file from structured data extracted from handwritten forms
        
        Args:
            data: Structured data containing form fields and tables
            
        Returns:
            Base64 encoded XLSX file
        """
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # Create a sheet for form fields
            self._add_form_fields_sheet(writer, data.get("form_fields", {}))
            
            # Create sheets for each table
            tables = data.get("tables", [])
            for i, table in enumerate(tables):
                self._add_table_sheet(writer, table, f"Table_{i+1}")
            
            # Apply styling to all worksheets
            for sheet_name in writer.sheets:
                self._apply_styling(writer.sheets[sheet_name])
        
        # Get the bytes and encode to base64
        excel_bytes = buffer.getvalue()
        return base64.b64encode(excel_bytes)
    
    def _add_form_fields_sheet(self, writer: pd.ExcelWriter, form_fields: Dict[str, Any]):
        """Add a sheet with form fields"""
        if not form_fields:
            # Create an empty form fields sheet
            pd.DataFrame({
                "Field": ["No form fields detected"],
                "Value": [""]
            }).to_excel(writer, sheet_name="Form Fields", index=False)
            return
        
        # Convert form fields to a DataFrame with Field and Value columns
        form_data = []
        for field, value in form_fields.items():
            form_data.append({"Field": field, "Value": value})
        
        # Create DataFrame and write to Excel
        df = pd.DataFrame(form_data)
        df.to_excel(writer, sheet_name="Form Fields", index=False)
    
    def _add_table_sheet(self, writer: pd.ExcelWriter, table_data: Dict[str, Any], sheet_name: str):
        """Add a sheet with table data"""
        # If table data has headers and rows
        headers = table_data.get("headers", [])
        rows = table_data.get("rows", [])
        
        if headers and rows:
            # Convert to DataFrame and write to Excel
            df = pd.DataFrame(rows, columns=headers)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        elif isinstance(table_data, list) and all(isinstance(item, dict) for item in table_data):
            # Handle case where table_data is a list of dictionaries
            df = pd.DataFrame(table_data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # Fallback for unstructured or unexpected table data
            pd.DataFrame({
                "Note": ["Table data could not be properly structured"]
            }).to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _apply_styling(self, worksheet):
        """Apply professional styling to the worksheet"""
        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Side(border_style="thin", color="DDDDDD")
        border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        # Apply header styling
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Apply row styling and auto-width
        row_count = worksheet.max_row
        col_count = worksheet.max_column
        
        # Auto-adjust column width
        for col_idx in range(1, col_count + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            
            for row_idx in range(1, row_count + 1):
                cell = worksheet[f"{column}{row_idx}"]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = max(max_length + 2, 10)  # Minimum width of 10
            worksheet.column_dimensions[column].width = min(adjusted_width, 40)  # Maximum width of 40
        
        # Apply alternating row colors and borders
        for row_idx in range(2, row_count + 1):
            # Alternating row background
            row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") if row_idx % 2 == 0 else None
            
            for col_idx in range(1, col_count + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = border
                
                if row_fill:
                    cell.fill = row_fill 